"""
this program converts an excel cells to an image according to inputed image
author: amin
"""

import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image
from tqdm import tqdm


def number_to_letters(n):
    """convert a decimal based number to alphabetical based number(A-ZZZ)"""
    result = ""
    while n > 0:
        n -= 1
        result = chr(ord("A") + n % 26) + result
        n //= 26
    return result


def set_excel_name(defualt_adress: str, excel_name: str) -> str:
    """repeated names will change"""
    if not excel_name:
        excel_name = re.sub(r"\.[a-zA-Z0-9]+$", "_colored.xlsx", defualt_adress)
    else:
        excel_name = re.sub(r"\.[a-zA-Z0-9]+$", "_colored.xlsx", excel_name)
    num = 0
    while os.path.isfile(excel_name):
        num += 1
        excel_name = re.sub(
            r"_colored\(?\d?\)?.xlsx", f"_colored({num}).xlsx", excel_name
        )
    return excel_name


def image_on_excel(
    image_adress: str, excel_adress: str = "", square_length: int = 10
) -> str:
    """it changes the cell colors acording to squares average color"""
    wb = Workbook()
    ws = wb.active
    img = Image.open(image_adress)
    w, h = img.size
    for col in tqdm(range(w // square_length)):
        for row in range(h // square_length):
            sr = 0
            sg = 0
            sb = 0
            for i in range((col - 1) * square_length, col * square_length):
                for j in range((row - 1) * square_length, row * square_length):
                    r, g, b = img.getpixel((i, j))
                    sr += r
                    sg += g
                    sb += b
            rgb_color = (
                sr // (square_length**2),
                sg // (square_length**2),
                sb // (square_length**2),
            )
            hex_color = f"{rgb_color[0]:02X}{rgb_color[1]:02X}{rgb_color[2]:02X}"
            color = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
            cell_id = f"{number_to_letters(col + 1)}{row + 1}"
            ws[cell_id].fill = color
        column = ws.column_dimensions[number_to_letters(col + 1)]
        column.width = 4

    excel_adress = set_excel_name(image_adress, excel_adress)
    wb.save(excel_adress)
    wb.close()
    output = "\n" + f"generated excel file saved as '{excel_adress}'".center(100, "*")
    return output


if __name__ == "__main__":
    square_scale = ""
    while not square_scale.isalnum():
        square_scale = input(
            "import a number for length of square that output from the picture\
(less number cause better quality): "
        )
    img_adr = input("paste image adress: ")
    excel_adr = input(
        "paste your excel adress and close it(if you want to make a new one just press \
Enter): "
    )
    print()
    if excel_adr:
        print(image_on_excel(img_adr, excel_adr, square_length=int(square_scale)))
    else:
        print(image_on_excel("sample.jpg", square_length=int(square_scale)))
